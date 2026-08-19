"""Portable Excel output template; no Node/Codex runtime dependency."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BLUE = "173F73"
LIGHT_BLUE = "EAF2FB"
GRADE_FILLS = {"A": "F8696B", "B": "F4B183", "C": "FFD966", "D": "9DC3E6", "E": "D9E1F2"}


def render_department_report_excel(payload: dict[str, Any], output_path: str | Path) -> Path:
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "报表概览"
    delayed = workbook.create_sheet(f"{payload['yesterday_date'][5:]}延期订单")
    important = workbook.create_sheet(f"{payload['yesterday_date'][5:]}重要订单")
    summary.merge_cells("A1:F1")
    summary["A1"] = f"{payload['department']} {payload['yesterday_date']}"
    summary["A1"].font = Font(bold=True, color="FFFFFF", size=18)
    summary["A1"].fill = PatternFill("solid", fgColor=BLUE)
    summary["A1"].alignment = Alignment(horizontal="center")
    metrics = [
        ("报告日期", payload["report_date"]), ("昨日日期", payload["yesterday_date"]),
        ("部门", payload["department"]), ("昨日计划订单", payload["yesterday_plan_order_count"]),
        ("昨日计划总公分数", payload["yesterday_plan_cm"]), ("已完工公分数", payload["completed_cm"]),
        ("未完工公分数", payload["unfinished_cm"]), ("公分数完成率", payload["completion_rate"]),
        ("完工/未完工订单", f"{payload['completed_order_count']}/{payload['unfinished_order_count']}"),
    ]
    for row_index, (label, value) in enumerate(metrics, start=3):
        summary.cell(row_index, 1, label)
        summary.cell(row_index, 2, value)
        summary.cell(row_index, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        summary.cell(row_index, 1).font = Font(bold=True, color=BLUE)
    summary["B10"].number_format = "0.0%"
    process_headers = ["工序", "计划公分数", "完工公分数", "未完工公分数"]
    _write_header(summary, 3, 4, process_headers)
    for row_index, row in enumerate(payload.get("process_stats", []), start=4):
        for column, value in enumerate((row["process"], row["planned_cm"], row["completed_cm"], row["unfinished_cm"]), start=4):
            summary.cell(row_index, column, value)
    headers = ["优先级", "订单子单号", "客户等级", "计划完成时间", "总公分数", "待完成工序", "交货日期", "工序状态"]
    limit = int(payload.get("example_limit", 3))
    _write_orders(delayed, headers, payload.get("delayed_example_orders", payload.get("delayed_open_orders", [])[:limit]))
    _write_orders(important, headers, payload.get("important_example_orders", payload.get("important_orders", [])[:limit]))
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2" if sheet is not summary else "A3"
        for column_index, column in enumerate(sheet.columns, start=1):
            letter = get_column_letter(column_index)
            sheet.column_dimensions[letter].width = min(36, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    workbook.save(path)
    return path


def _write_header(sheet, row: int, column: int, headers: list[str]) -> None:
    for offset, label in enumerate(headers):
        cell = sheet.cell(row, column + offset, label)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True, color="FFFFFF")


def _write_orders(sheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    _write_header(sheet, 1, 1, headers)
    for row_index, row in enumerate(rows, start=2):
        values = [
            row.get("priority", ""), row.get("product_order_no", ""), row.get("customer_grade", ""),
            row.get("planned_date", ""), float(row.get("centimeters") or 0), row.get("incomplete_processes", ""),
            row.get("delivery_date", ""), row.get("statuses", ""),
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row_index, column, value)
        grade = str(row.get("customer_grade", "")).strip().upper()
        if grade in GRADE_FILLS:
            sheet.cell(row_index, 3).fill = PatternFill("solid", fgColor=GRADE_FILLS[grade])
